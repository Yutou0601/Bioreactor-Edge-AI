"""
實驗報表匯出
============
把批次（experiment_runs）產生的量測結果，輸出成洪博熟悉的批次表格式：
  - Excel（.xlsx，綠底表格，可直接交付）
  - CSV（純文字，供後續分析）

只用可信訊號的量測結果；CH4% 欄標明「參考」。
"""

import csv
import io

# 報表欄位（對應 2026-07-22 定案協定）。前 6 欄＝設定條件，其餘＝量測結果。
COLUMNS = [
    ("run_id",            "批次實驗"),
    ("gas_ratio",         "氣體比例(H2:CO2)"),
    ("n_minutes",         "循環時間(每時幾分)"),
    ("intake_band",       "自動補氣(下限→上限)"),
    ("baseline",          "基準(CH4%/CO2%/壓力)"),
    ("target_hours",      "預計時長(hr)"),
    ("total_hours",       "實際總時間(hr)"),
    ("cycle_count",       "循環數(完整/全部)"),
    ("gap_note",          "記錄中斷"),
    ("drop_rate_median",  "下降速率中位數(kg/cm²/hr)"),
    ("drop_rate_spread",  "下降速率離散(IQR·範圍)"),
    ("culture_drift",     "進氣前ORP漂移(末-首)"),
    ("vent_orp",          "排氣ORP峰值(mV)"),
    ("vent_ph",           "排氣pH峰值"),
    ("vent_co2",          "排氣CO2%峰值"),
    ("vent_ch4_peak_ref", "排氣CH4%峰值"),
    ("peaks_manual_note", "峰值來源"),
    ("status",            "狀態"),
]
SETTING_COLS = 6   # 前幾欄為設定條件（灰底），其餘為量測結果（綠底）


def _flatten(run: dict) -> dict:
    """把 run + run['results'] 攤平成單層 dict，並組合衍生欄位。"""
    flat = {k: v for k, v in run.items() if k != "results"}
    flat.update(run.get("results", {}))
    flat["status"] = {"planned": "已規劃", "running": "進行中", "done": "已完成"}.get(
        flat.get("status"), flat.get("status", ""))
    # 衍生欄位
    flat["intake_band"] = f"{run.get('intake_lower')}→{run.get('intake_upper')}"
    flat["baseline"] = f"{run.get('baseline_ch4')}/{run.get('baseline_co2')}/{run.get('baseline_pressure')}"
    # 完整週期／總週期分開寫，讓「這批有多少可用資料」在表上一眼看得到
    flat["cycle_count"] = f"{flat.get('n_cycles_complete', 0)} / {flat.get('n_cycles', 0)}"
    n_gaps = flat.get("n_gaps") or 0
    flat["gap_note"] = f"{n_gaps} 次 / {flat.get('gap_hours') or 0} hr" if n_gaps else "無"
    # 離散度：投稿時中位數需附散布。IQR 需 ≥2 完整循環才算得出來，否則只列範圍
    iqr, lo, hi = flat.get("drop_rate_iqr"), flat.get("drop_rate_min"), flat.get("drop_rate_max")
    if iqr is not None:
        flat["drop_rate_spread"] = f"IQR {iqr} ({lo}–{hi})"
    elif lo is not None:
        flat["drop_rate_spread"] = f"n=1 ({lo})"
    else:
        flat["drop_rate_spread"] = ""
    # 峰值來源：標明哪些是現場手動輸入（較準）、其餘為感測器自動抓（1/min 可能錯過峰）
    manual = flat.get("peaks_manual") or []
    LBL = {"orp": "ORP", "ph": "pH", "co2": "CO2", "ch4": "CH4"}
    flat["peaks_manual_note"] = ("手動:" + "/".join(LBL[k] for k in manual if k in LBL)) \
        if manual else "全自動"
    return flat


def _cell(flat: dict, key: str):
    v = flat.get(key)
    if v is None:
        return ""
    if key == "n_minutes":
        return f"{v:g} 分"
    return v


def to_csv(runs: list) -> str:
    """輸出 CSV 字串（utf-8-sig 由呼叫端加 BOM）。"""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([zh for _, zh in COLUMNS])
    for run in runs:
        flat = _flatten(run)
        w.writerow([_cell(flat, key) for key, _ in COLUMNS])
    return buf.getvalue()


# ── 每循環特徵表（餵模型用；每列＝一個循環週期）──────────
CYCLE_COLUMNS = [
    ("run_id",            "批次"),
    ("n_minutes",         "循環時間(每時幾分)"),
    ("cycle",             "週期序"),
    ("start",             "起始時間"),
    ("duration_hr",       "時長(hr)"),
    ("pressure_start",    "壓力起"),
    ("pressure_end",      "壓力末"),
    ("drop_rate",         "下降速率(kg/cm²/hr)"),
    ("slope_early",       "早段斜率"),
    ("slope_late",        "晚段斜率"),
    ("flattening",        "平緩化(早-晚·疑產甲烷)"),
    ("pre_injection_orp", "進氣前ORP(菌群共變數)"),
    ("orp_crash",         "ORP崩落"),
    ("quality",           "資料完整性"),
]


def _cycle_cell(row: dict, key: str):
    v = row.get(key)
    if v is None:
        return ""
    if key == "n_minutes":
        return f"{v:g} 分"
    return v


def cycles_to_csv(cycle_rows: list) -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([zh for _, zh in CYCLE_COLUMNS])
    for row in cycle_rows:
        w.writerow([_cycle_cell(row, key) for key, _ in CYCLE_COLUMNS])
    return buf.getvalue()


def cycles_to_xlsx_bytes(cycle_rows: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    GREEN = PatternFill("solid", fgColor="92D050")
    COV = PatternFill("solid", fgColor="FCE4A0")   # 共變數欄位淡黃強調
    HFONT = Font(name="微軟正黑體", size=11, bold=True, color="000000")
    BFONT = Font(name="微軟正黑體", size=11)
    NOTE = Font(name="微軟正黑體", size=9, italic=True, color="595959")
    THIN = Side(style="thin", color="000000")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "每循環特徵"
    ws.cell(1, 1, "生物甲烷化 — 每循環特徵表（餵模型用）").font = Font(name="微軟正黑體", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CYCLE_COLUMNS))

    hr = 3
    for j, (key, zh) in enumerate(CYCLE_COLUMNS, 1):
        c = ws.cell(hr, j, zh)
        c.font = HFONT
        c.fill = COV if key == "pre_injection_orp" else GREEN
        c.alignment = CEN
        c.border = BORDER
    WARN = PatternFill("solid", fgColor="FFC7CE")   # 非完整週期標紅，避免誤用
    for i, row in enumerate(cycle_rows):
        incomplete = not row.get("complete")
        for j, (key, _) in enumerate(CYCLE_COLUMNS, 1):
            c = ws.cell(hr + 1 + i, j, _cycle_cell(row, key))
            c.font = BFONT
            c.alignment = CEN
            c.border = BORDER
            if incomplete and key == "quality":
                c.fill = WARN

    for j, w in enumerate([8, 15, 8, 18, 10, 9, 9, 18, 10, 10, 16, 18, 10, 14], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[hr].height = 34

    note_row = hr + 1 + len(cycle_rows) + 1
    ws.cell(note_row, 1, "※ 每列＝一個補氣循環。下降速率為反應變數、進氣前ORP為菌群成熟度共變數、"
            "循環時間(n)為控制因子。分析時用進氣前ORP扣除菌群漂移。"
            "※ 建模只取「資料完整性＝完整」之列；標紅者起點或終點的補氣未被觀測到"
            "（多為記錄中斷），其斜率平緩化不予計算。").font = NOTE
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(CYCLE_COLUMNS))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def to_xlsx_bytes(runs: list) -> bytes:
    """輸出 .xlsx 位元組（綠底表格，比照洪博格式）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    GREEN = PatternFill("solid", fgColor="92D050")
    GREY = PatternFill("solid", fgColor="F2F2F2")
    HFONT = Font(name="微軟正黑體", size=11, bold=True, color="000000")
    BFONT = Font(name="微軟正黑體", size=11)
    NOTE = Font(name="微軟正黑體", size=9, italic=True, color="595959")
    THIN = Side(style="thin", color="000000")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "實驗批次結果"

    ws.cell(1, 1, "生物甲烷化 — 循環時間實驗批次結果").font = Font(name="微軟正黑體", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    hr = 3
    for j, (_, zh) in enumerate(COLUMNS, 1):
        c = ws.cell(hr, j, zh)
        c.font = HFONT
        c.fill = GREY if j <= SETTING_COLS else GREEN   # 前 N 欄=設定條件(灰), 其餘=量測結果(綠)
        c.alignment = CEN
        c.border = BORDER

    for i, run in enumerate(runs):
        flat = _flatten(run)
        for j, (key, _) in enumerate(COLUMNS, 1):
            c = ws.cell(hr + 1 + i, j, _cell(flat, key))
            c.font = BFONT
            c.alignment = CEN
            c.border = BORDER

    widths = [10, 13, 13, 16, 18, 11, 12, 16, 14, 18, 20, 16, 12, 10, 11, 12, 14, 9]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[hr].height = 40

    note_row = hr + 1 + len(runs) + 1
    ws.cell(note_row, 1, "※ 灰底＝設定條件，綠底＝量測結果（由感測訊號自動計算）。"
            "「進氣前ORP漂移」為菌群成熟度共變數，用於分析時扣除菌群漂移。"
            "CH4% 僅取排氣峰值當參考，不作為證據。"
            "※ 下降速率中位數只採計完整循環；「記錄中斷」非零時，該批次期間反應器仍在運轉，"
            "僅資料未被記錄，中斷期間的補氣無法還原。").font = NOTE
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(COLUMNS))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
